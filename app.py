# -*- coding: utf-8 -*-

import io
from flask import Flask, request, jsonify, send_file

# Importaciones de openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import Color
from openpyxl.cell import MergedCell
from openpyxl.formatting.rule import ColorScaleRule, IconSetRule, DataBarRule

# --- Inicialización de la Aplicación Flask ---
app = Flask(__name__)

# === FUNCIONES DE AYUDA PARA EXTRAER DATOS ===

def get_serializable_color(color_obj):
    """Convierte un objeto Color de openpyxl a un string hexadecimal (ARGB)."""
    if color_obj and isinstance(color_obj, Color) and color_obj.rgb:
        return color_obj.rgb
    # Podrías añadir lógica para otros tipos de color si es necesario (theme, indexed)
    return None

def extract_styles_from_cell(cell):
    """Extrae estilos de una celda y los devuelve en un formato serializable."""
    style_data = {}
    if not cell.has_style:
        return style_data

    # Fuente (Font)
    if cell.font:
        font_data = {
            'name': cell.font.name,
            'sz': cell.font.sz,
            'bold': cell.font.bold,
            'italic': cell.font.italic,
            'color': get_serializable_color(cell.font.color)
        }
        # Eliminar claves con valor None o False para un JSON más limpio
        style_data['font'] = {k: v for k, v in font_data.items() if v}

    # Relleno (Fill)
    if cell.fill and cell.fill.fill_type:
        fill_data = {
            'pattern': cell.fill.fill_type,
            'start_color': get_serializable_color(cell.fill.start_color),
            'end_color': get_serializable_color(cell.fill.end_color)
        }
        style_data['fill'] = {k: v for k, v in fill_data.items() if v}
        
    # Bordes (Border)
    if cell.border:
        def get_side_style(side):
            if side and side.style:
                return {'style': side.style, 'color': get_serializable_color(side.color)}
            return None
        
        border_data = {
            'left': get_side_style(cell.border.left),
            'right': get_side_style(cell.border.right),
            'top': get_side_style(cell.border.top),
            'bottom': get_side_style(cell.border.bottom)
        }
        style_data['border'] = {k: v for k, v in border_data.items() if v}
        
    # Alineación (Alignment)
    if cell.alignment:
        alignment_data = {
            'horizontal': cell.alignment.horizontal,
            'vertical': cell.alignment.vertical,
            'wrap_text': cell.alignment.wrap_text
        }
        style_data['alignment'] = {k: v for k, v in alignment_data.items() if v}
        
    if cell.number_format and cell.number_format != 'General':
        style_data['numFmt'] = cell.number_format
        
    return style_data

def extract_conditional_formats(ws):
    """Extrae todas las reglas de formato condicional de una hoja."""
    formats_data = []
    # ws.conditional_formatting es un diccionario donde la clave es el rango
    for range_string, rules in ws.conditional_formatting.items():
        for rule in rules:
            rule_info = {
                'range': range_string,
                'type': rule.type
            }
            # Extraer detalles específicos según el tipo de regla
            if isinstance(rule, ColorScaleRule) and rule.colorScale:
                rule_info['color_scale'] = {
                    'colors': [get_serializable_color(c) for c in rule.colorScale.color],
                    'values': [cfvo.val for cfvo in rule.colorScale.cfvo]
                }
            elif isinstance(rule, DataBarRule) and rule.dataBar:
                 rule_info['data_bar'] = {
                    'color': get_serializable_color(rule.dataBar.color),
                    'min_length': rule.dataBar.minLength,
                    'max_length': rule.dataBar.maxLength,
                 }
            elif hasattr(rule, 'formula') and rule.formula:
                 rule_info['formula'] = rule.formula
            
            formats_data.append(rule_info)
    return formats_data

def extract_charts(ws):
    """Extrae toda la información de los gráficos de una hoja."""
    charts_data = []
    # Los gráficos se almacenan en el atributo (semi-privado) _charts
    for chart in ws._charts:
        chart_info = {
            'type': chart.__class__.__name__, # Ej: 'BarChart', 'PieChart'
            'title': chart.title.text.text if chart.title and chart.title.text else None,
            'anchor': str(chart.anchor), # Dónde está ubicado el gráfico
            'series': []
        }
        
        # Extraer las series de datos
        for s in chart.series:
            series_info = {
                'header': s.tx.v if s.tx else None,
                # La referencia a los valores (ej: 'Hoja1!$F$2:$F$7')
                'values': str(s.val.ref) if s.val and hasattr(s.val, 'ref') else None,
                # La referencia a las categorías/etiquetas (ej: 'Hoja1!$A$2:$A$7')
                'categories': str(s.cat.ref) if s.cat and hasattr(s.cat, 'ref') else None,
            }
            chart_info['series'].append(series_info)
        
        charts_data.append(chart_info)
    return charts_data


# --- ENDPOINT DE CHEQUEO DE SALUD ---
@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({"status": "ok"}), 200

# --- ENDPOINT PARA ANALIZAR EXCEL ---
@app.route('/parse-excel', methods=['POST'])
def parse_excel():
    if 'excel_file' not in request.files:
        return jsonify({"error": "No se encontró el archivo en la petición (se esperaba el campo 'excel_file')."}), 400
    
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({"error": "No se seleccionó ningún archivo."}), 400
    
    try:
        in_memory_file = io.BytesIO(file.read())
        # data_only=True lee el valor de las fórmulas, no las fórmulas en sí.
        # Si necesitas las fórmulas, ponlo en False.
        wb = load_workbook(filename=in_memory_file, data_only=True)
        
        parsed_data = {'sheets': []}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = {
                'name': sheet_name,
                'merged_cells': [str(r) for r in ws.merged_cells.ranges],
                'data': [],
                # --- NUEVAS SECCIONES ---
                'conditional_formats': extract_conditional_formats(ws),
                'charts': extract_charts(ws)
            }
            
            rows_data = []
            for row in ws.iter_rows():
                row_list = []
                for cell in row:
                    cell_info = {'address': cell.coordinate, 'value': cell.value}
                    if isinstance(cell, MergedCell):
                        cell_info['is_merged_part'] = True
                    else:
                        cell_info['style'] = extract_styles_from_cell(cell)
                    row_list.append(cell_info)
                rows_data.append(row_list)
            
            sheet_data['data'] = rows_data
            parsed_data['sheets'].append(sheet_data)
            
        return jsonify(parsed_data)
    
    except Exception as e:
        # Imprime el error en la consola del servidor para depuración
        import traceback
        print(f"Error en /parse-excel: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Error interno al procesar el archivo Excel: {str(e)}"}), 500

# --- Punto de Entrada de la Aplicación ---
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
